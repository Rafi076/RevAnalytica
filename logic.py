from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib
from collections import Counter

# Path to the Excel file
file = pathlib.Path("Sales_Data_Record.xlsx")

# Function to get the most frequent sold item
def get_frequent_item(frequentItemEntry):
    if not file.exists():
        messagebox.showerror("Error", "Excel file not found.")
        return

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        product_names = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if row[5]:  # Ensure the "Sold Product Name" column (index 5) is not None
                product_names.append(row[5])

        if product_names:
            # Count the frequency of each product
            product_counter = Counter(product_names)
            # Get the two most common products
            top_two_products = product_counter.most_common(4)  # Returns a list of tuples [(product1, count1), (product2, count2)]

            # Extract just the product names
            most_frequent_items = [item[0] for item in top_two_products]

            # Display the top two products in the entry field
            frequentItemEntry.delete(0, "end")
            frequentItemEntry.insert(0, ", ".join(most_frequent_items))
        else:
            messagebox.showinfo("Info", "No transactions found.")
        
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def search_transaction_details(transactionIDEntry, customerIDEntry, salespersonIDEntry, dateEntry, monthEntry, itemSetEntry,
                               productCategoryEntry, quantitySoldEntry, pricePerUnitEntry, totalSaleAmountEntry,
                               paymentMethodEntry, profitMarginEntry, membershipEntry):
    search_transaction_id = transactionIDEntry.get().strip()

    if not search_transaction_id:
        messagebox.showerror("Error", "Transaction ID is required to search.")
        return

    if not file.exists():
        messagebox.showerror("Error", "Excel file not found.")
        return

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        found = False
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            print(row)  # Debugging: Check row data
            if str(row[0]) == search_transaction_id:  # Ensure matching as string
                # Clear previous entries
                for entry in [transactionIDEntry, customerIDEntry, salespersonIDEntry, dateEntry, monthEntry,
                              itemSetEntry, productCategoryEntry, quantitySoldEntry, pricePerUnitEntry,
                              totalSaleAmountEntry, paymentMethodEntry, profitMarginEntry, membershipEntry]:
                    entry.delete(0, "end")

                # Populate fields with data
                transactionIDEntry.insert(0, row[0] if row[0] else "")
                customerIDEntry.insert(0, row[1] if row[1] else "")
                salespersonIDEntry.insert(0, row[2] if row[2] else "")
                dateEntry.insert(0, row[3] if row[3] else "")
                monthEntry.insert(0, row[4] if row[4] else "")
                itemSetEntry.insert(0, row[5] if row[5] else "")
                productCategoryEntry.insert(0, row[6] if row[6] else "")
                quantitySoldEntry.insert(0, row[7] if row[7] else "")
                pricePerUnitEntry.insert(0, row[8] if row[8] else "")
                totalSaleAmountEntry.insert(0, row[9] if row[9] else "")
                paymentMethodEntry.insert(0, row[10] if row[10] else "")
                profitMarginEntry.insert(0, row[11] if row[11] else "")
                membershipEntry.insert(0, row[12] if row[12] else "")

                found = True
                break

        if not found:
            messagebox.showinfo("Not Found", "Transaction ID not found.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Refresh Transaction Fields
def refresh_transaction_fields(entries):
    """
    Clears all the input fields provided in the entries list.
    """
    for entry in entries:
        entry.delete(0, "end")  # Clear the content of each entry widget
    if entries:  # Ensure at least one entry exists
        entries[0].focus()  # Set focus to the first entry widget (Transaction ID)


def submit_transaction_details(entries):
    # Extract and clean data from entry fields
    transaction_id, customer_id, salesperson_id, date, month_of_sale, item_set, product_category, \
    quantity_sold, price_per_unit, total_sale_amount, payment_method, profit_margin, membership_status = [
        entry.get().strip() for entry in entries
    ]

    # Validation checks
    if not transaction_id:
        messagebox.showerror("Error", "Transaction ID is required.")
        return
    if not customer_id:
        messagebox.showerror("Error", "Customer ID is required.")
        return
    if not salesperson_id:
        messagebox.showerror("Error", "Salesperson ID is required.")
        return
    if not date:
        messagebox.showerror("Error", "Date is required.")
        return
    if not month_of_sale:
        messagebox.showerror("Error", "Month of Sale is required.")
        return

    # Validate Quantity Sold
    cleaned_quantity = quantity_sold.replace(",", "").strip()
    if not cleaned_quantity.isdigit() or int(cleaned_quantity) <= 0:
        # Check if it's a valid bit string
        if not all(c in "01" for c in cleaned_quantity):
            messagebox.showerror("Error", "Quantity Sold must be a positive integer or a valid bit string.")
            return

    # Validate Price Per Unit
    try:
        cleaned_price = price_per_unit.replace(",", "").strip()
        # Allow bit string format or numeric validation
        if not all(c in "01" for c in cleaned_price):  # Check for bit string
            cleaned_price = float(cleaned_price)  # If not bit string, must be numeric
            if cleaned_price <= 0:
                raise ValueError
    except ValueError:
        messagebox.showerror("Error", "Price Per Unit must be a positive number or a valid bit string.")
        return

    # Validate Total Sale Amount (Positive Number)
    try:
        cleaned_total = float(total_sale_amount.replace(",", "").strip())
        if cleaned_total <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror("Error", "Total Sale Amount must be a positive number.")
        return

    # Validate Profit Margin (Number)
    try:
        float(profit_margin.replace(",", "").strip())  # Allow any numeric value
    except ValueError:
        messagebox.showerror("Error", "Profit Margin must be a valid number.")
        return

    try:
        # Ensure the file exists or create it
        if not file.exists():
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([
                "Transaction ID", "Customer ID", "Salesperson ID", "Date", "Month of Sale", "Item Set",
                "Product Category", "Quantity Sold", "Price Per Unit", "Total Sale Amount",
                "Payment Method", "Profit Margin", "Membership Status"
            ])
            workbook.save(file)

        # Load workbook and append data
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        sheet.append([
            transaction_id, customer_id, salesperson_id, date, month_of_sale, item_set,
            product_category, cleaned_quantity, cleaned_price, cleaned_total,
            payment_method, profit_margin, membership_status
        ])
        workbook.save(file)

        # Success message and clear fields
        messagebox.showinfo("Success", "Transaction details saved successfully!")
        refresh_transaction_fields(entries)

    except PermissionError:
        messagebox.showerror("Error", "The file is open in another program. Please close it and try again.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


